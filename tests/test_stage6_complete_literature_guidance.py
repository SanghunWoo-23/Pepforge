import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "spps_planner_app"))
os.environ.setdefault("SPPS_PLANNER_USER_DATA", "/tmp/pepforge_stage6_spps_data")

from spps_planner.engine import PlanInput, recommend_cleavage_preset, validate_plan
from spps_planner.export import export_csvs, export_excel
from spps_planner.literature_guidance import generate_literature_guidance


def _categories(frame):
    return set(frame["category"].astype(str))


def test_confirmed_ac_eemqrr_cleavage_contract_is_preserved():
    inp = PlanInput(sequence="Ac-EEMQRR-NH2", resin="Rink Amide AM", scale_mmol=1)
    rec = recommend_cleavage_preset(inp)
    assert rec["preset"] == "DEFAULT_TFA_WATER"
    assert "30 eq" in rec["reason"] and "95% TFA / 5% DW" in rec["reason"] and "no TIS" in rec["reason"]


def test_complete_guidance_separates_major_decision_domains():
    inp = PlanInput(sequence="QDGVIITFWYNNQHHCCC-NH2", resin="Rink Amide AM", scale_mmol=1)
    frame = generate_literature_guidance(inp)
    required = {
        "workflow", "coupling", "protecting_group", "resin_linker", "cleavage",
        "aspartimide", "difficult_sequence", "disulfide_cyclization",
        "oxidation_options", "chemical_liability", "workup_counterion",
        "structure_validation", "external_validation", "sustainability",
    }
    assert required.issubset(_categories(frame))
    assert set(frame.columns) == {
        "category", "trigger", "priority", "recommendation",
        "verification_required", "evidence", "limitation",
    }
    assert frame["evidence"].astype(str).str.contains("DOI|PMID").all()


def test_resin_routes_distinguish_mild_cleavage_from_global_deprotection():
    ctc = generate_literature_guidance(PlanInput(sequence="ACDC", resin="2-CTC"))
    ctc_text = " ".join(ctc.loc[ctc.category == "resin_linker", "recommendation"])
    assert "mild protected-fragment cleavage" in ctc_text
    assert "not interchangeable" in ctc_text
    rink = generate_literature_guidance(PlanInput(sequence="ACDC-NH2", resin="Rink Amide AM"))
    rink_text = " ".join(rink.loc[rink.category == "resin_linker", "recommendation"])
    assert "C-terminal amide" in rink_text and "global-cleavage" in rink_text


def test_aspartimide_guidance_is_validation_bounded():
    frame = generate_literature_guidance(PlanInput(sequence="ADGDSN-NH2"))
    row = frame.loc[frame.category == "aspartimide"].iloc[0]
    assert "1 M Oxyma" in row.recommendation
    assert "not a universal default" in row.recommendation
    assert "HPLC/MS" in row.verification_required


def test_noncanonical_units_never_inherit_canonical_parameters():
    frame = generate_literature_guidance(PlanInput(sequence="Ac-bAla-gAla-dK-NH2"))
    row = frame.loc[frame.category == "noncanonical_backbone"].iloc[0]
    assert "Do not inherit canonical-alpha residue parameters" in row.recommendation
    assert "dedicated beta/gamma" in row.verification_required


def test_explicit_protecting_groups_and_met_route_are_preserved():
    frame = generate_literature_guidance(PlanInput(sequence="C(Acm)M-R(Pbf)-NH2"))
    pg_text = " ".join(frame.loc[frame.category == "protecting_group", "trigger"])
    assert "Acm" in pg_text and "Pbf" in pg_text
    rec = recommend_cleavage_preset(PlanInput(sequence="AMAA-NH2"))
    assert rec["preset"] == "REAGENT_H"
    cys_rows = frame.loc[(frame.category == "cleavage") & (frame.trigger == "Cys present")]
    assert "+56.0626 Da" in cys_rows.iloc[0].verification_required


def test_guidance_is_integrated_into_validation_and_exports(tmp_path):
    inp = PlanInput(sequence="QDGVIITFWYNNQHHCCC-NH2", resin="Rink Amide AM", scale_mmol=0.1)
    validation = validate_plan(inp)
    assert validation["area"].astype(str).str.startswith("literature/").any()
    csv_dir = tmp_path / "csv"
    export_csvs(inp, csv_dir)
    assert (csv_dir / "literature_guidance.csv").exists()
    xlsx = tmp_path / "plan.xlsx"
    export_excel(inp, xlsx)
    from openpyxl import load_workbook
    assert "11_LITERATURE_GUIDANCE" in load_workbook(xlsx, read_only=True).sheetnames
