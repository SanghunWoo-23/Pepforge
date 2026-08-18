from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .engine import PlanInput, generate_step_matrix, generate_detailed_operations, generate_materials, generate_step_reagent_plan, generate_step_materials, generate_ml_ready_log, generate_excel_like_synthesis_table, generate_printable_checklist, generate_cleavage_cocktail, cleavage_cocktail_presets, validate_plan, plan_summary
from .literature_guidance import generate_literature_guidance
from .display import total_materials_display, concise_plan

HEADER_FILL = "E5E7EB"
TITLE_FILL = "1F2937"
BORDER = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for j, col in enumerate(df.columns, start_col):
        cell = ws.cell(start_row, j, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, value in enumerate(row, start_col):
            cell = ws.cell(i, j, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    for j, col in enumerate(df.columns, start_col):
        series = df[col].head(100).astype(object).where(pd.notna(df[col].head(100)), "")
        values = [len(str(col))] + [len(str(v)) for v in series]
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(55, max(values) + 2))


def export_excel(inp: PlanInput, out_path: str | Path):
    out_path = Path(out_path)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "00_EXCEL_LIKE_PLAN"
    _write_df(ws0, concise_plan(generate_excel_like_synthesis_table(inp)), 1, 1)
    ws = wb.create_sheet("01_SUMMARY")
    _write_df(ws, pd.DataFrame([plan_summary(inp)]), 1, 1)
    ws2 = wb.create_sheet("02_PRINT_CHECKLIST")
    _write_df(ws2, generate_printable_checklist(inp), 1, 1)
    ws2b = wb.create_sheet("03_SYNTHESIS_FORM")
    _write_df(ws2b, generate_detailed_operations(inp), 1, 1)
    ws3 = wb.create_sheet("04_SELECTED_MATERIALS_STEP")
    _write_df(ws3, generate_step_materials(inp), 1, 1)
    ws3b = wb.create_sheet("04B_RAW_MATERIAL_TOTAL")
    _write_df(ws3b, total_materials_display(generate_materials(inp), inp.resin), 1, 1)
    ws4 = wb.create_sheet("05_STEP_MATRIX")
    _write_df(ws4, generate_step_matrix(inp), 1, 1)
    ws5 = wb.create_sheet("06_STEP_REAGENTS")
    _write_df(ws5, generate_step_reagent_plan(inp), 1, 1)
    ws6 = wb.create_sheet("07_ML_READY_LOG")
    _write_df(ws6, generate_ml_ready_log(inp), 1, 1)
    ws7 = wb.create_sheet("08_VALIDATION")
    _write_df(ws7, validate_plan(inp), 1, 1)
    ws8 = wb.create_sheet("09_CLEAVAGE_COCKTAIL")
    _write_df(ws8, generate_cleavage_cocktail(inp), 1, 1)
    ws9 = wb.create_sheet("10_CLEAVAGE_PRESETS")
    _write_df(ws9, cleavage_cocktail_presets(), 1, 1)
    ws10 = wb.create_sheet("11_LITERATURE_GUIDANCE")
    _write_df(ws10, generate_literature_guidance(inp), 1, 1)
    for sh in wb.worksheets:
        sh.freeze_panes = "A2"
    wb.save(out_path)


def export_csvs(inp: PlanInput, out_dir: str | Path):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    concise_plan(generate_excel_like_synthesis_table(inp)).to_csv(out_dir / "excel_like_synthesis_plan.csv", index=False, encoding="utf-8-sig")
    generate_step_matrix(inp).to_csv(out_dir / "step_matrix.csv", index=False, encoding="utf-8-sig")
    generate_detailed_operations(inp).to_csv(out_dir / "synthesis_form_wash_by_wash.csv", index=False, encoding="utf-8-sig")
    generate_step_materials(inp).to_csv(out_dir / "selected_materials_stepwise.csv", index=False, encoding="utf-8-sig")
    generate_materials(inp).to_csv(out_dir / "raw_material_use.csv", index=False, encoding="utf-8-sig")
    total_materials_display(generate_materials(inp), inp.resin).to_csv(out_dir / "selected_total_materials_visible.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame([plan_summary(inp)]).to_csv(out_dir / "summary.csv", index=False, encoding="utf-8-sig")
    generate_step_reagent_plan(inp).to_csv(out_dir / "step_reagent_plan.csv", index=False, encoding="utf-8-sig")
    generate_ml_ready_log(inp).to_csv(out_dir / "spps_ml_ready_log.csv", index=False, encoding="utf-8-sig")
    generate_printable_checklist(inp).to_csv(out_dir / "printable_synthesis_checklist.csv", index=False, encoding="utf-8-sig")
    validate_plan(inp).to_csv(out_dir / "validation_warnings.csv", index=False, encoding="utf-8-sig")
    generate_cleavage_cocktail(inp).to_csv(out_dir / "cleavage_cocktail.csv", index=False, encoding="utf-8-sig")
    cleavage_cocktail_presets().to_csv(out_dir / "cleavage_cocktail_presets.csv", index=False, encoding="utf-8-sig")
    generate_literature_guidance(inp).to_csv(out_dir / "literature_guidance.csv", index=False, encoding="utf-8-sig")
