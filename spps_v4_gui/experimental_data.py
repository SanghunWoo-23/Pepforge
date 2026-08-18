"""Persistent experimental knowledge base for SPPS Planner V4.0.0.

The V4 layer is additive: it does not replace planner calculations.  It stores
real loading/cleavage observations, preserves raw source text, and separates
parsed records from operator-verified records before supervised training.
"""
from __future__ import annotations

from datetime import datetime, timezone
import csv
import json
import math
from pathlib import Path
import re
import sqlite3
import tempfile
from typing import Any, Iterable, Mapping
from uuid import uuid4
import zipfile


SCHEMA_VERSION = 1
STATUSES = ("parsed", "verified", "incomplete", "excluded")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _id() -> str:
    return uuid4().hex


def default_db_path() -> Path:
    try:
        from spps_planner.user_paths import user_file
        return Path(user_file("experimental_v4.sqlite"))
    except Exception:
        return Path.home() / ".spps_planner" / "data" / "experimental_v4.sqlite"


def _connect(path: str | Path | None = None) -> sqlite3.Connection:
    destination = Path(path) if path else default_db_path()
    destination.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(destination)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    return con


def initialize(path: str | Path | None = None) -> Path:
    destination = Path(path) if path else default_db_path()
    with _connect(destination) as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS import_sources (
                source_id TEXT PRIMARY KEY,
                imported_at TEXT NOT NULL,
                source_name TEXT NOT NULL,
                source_kind TEXT NOT NULL,
                source_path TEXT NOT NULL,
                sha256 TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS loading_records (
                record_id TEXT PRIMARY KEY,
                status TEXT NOT NULL,
                date TEXT NOT NULL DEFAULT '',
                resin_type TEXT NOT NULL DEFAULT '',
                resin_note TEXT NOT NULL DEFAULT '',
                amino_acid_raw TEXT NOT NULL DEFAULT '',
                amino_acid_normalized TEXT NOT NULL DEFAULT '',
                stereochemistry TEXT NOT NULL DEFAULT '',
                protecting_group TEXT NOT NULL DEFAULT '',
                aa_eq REAL,
                base TEXT NOT NULL DEFAULT '',
                base_eq REAL,
                coupling_reagent TEXT NOT NULL DEFAULT '',
                coupling_reagent_eq REAL,
                additive TEXT NOT NULL DEFAULT '',
                additive_eq REAL,
                loading_time_h REAL,
                loading_solvent TEXT NOT NULL DEFAULT '',
                capping_performed INTEGER,
                capping_method TEXT NOT NULL DEFAULT '',
                resin_sample_weight_mg REAL,
                absorbance REAL,
                loading_rate_mmol_g REAL,
                raw_note TEXT NOT NULL DEFAULT '',
                source_id TEXT,
                source_locator TEXT NOT NULL DEFAULT '',
                outlier_flag INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(source_id) REFERENCES import_sources(source_id)
            );

            CREATE TABLE IF NOT EXISTS cleavage_records (
                record_id TEXT PRIMARY KEY,
                status TEXT NOT NULL,
                product TEXT NOT NULL DEFAULT '',
                sequence TEXT NOT NULL DEFAULT '',
                scale_mmol REAL,
                operator TEXT NOT NULL DEFAULT '',
                tfa_ml REAL,
                tis_ml REAL,
                water_ml REAL,
                other_scavengers_json TEXT NOT NULL DEFAULT '{}',
                cleavage_eq REAL,
                cleavage_time_h REAL,
                temperature_c REAL,
                ether_ml REAL,
                ether_ratio TEXT NOT NULL DEFAULT '',
                filter_ether_ml REAL,
                filter_speed TEXT NOT NULL DEFAULT '',
                crude_g REAL,
                precipitation_good INTEGER,
                separation_problem INTEGER,
                concentration_recommended INTEGER,
                remove_tis_recommended INTEGER,
                overnight_hardening INTEGER,
                raw_observation TEXT NOT NULL DEFAULT '',
                raw_filter_note TEXT NOT NULL DEFAULT '',
                source_id TEXT,
                source_locator TEXT NOT NULL DEFAULT '',
                outlier_flag INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(source_id) REFERENCES import_sources(source_id)
            );

            CREATE INDEX IF NOT EXISTS idx_loading_lookup
              ON loading_records(resin_type, amino_acid_normalized, status);
            CREATE INDEX IF NOT EXISTS idx_cleavage_lookup
              ON cleavage_records(product, status);
            """
        )
    return destination


def _sha256(path: Path) -> str:
    import hashlib
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _register_source(con: sqlite3.Connection, path: Path, kind: str, note: str = "") -> str:
    digest = _sha256(path) if path.is_file() else ""
    existing = con.execute(
        "SELECT source_id FROM import_sources WHERE sha256=? AND source_kind=?",
        (digest, kind),
    ).fetchone() if digest else None
    if existing:
        return str(existing["source_id"])
    source_id = _id()
    con.execute(
        "INSERT INTO import_sources VALUES (?,?,?,?,?,?,?)",
        (source_id, _now(), path.name, kind, str(path), digest, note),
    )
    return source_id


def _float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "—", "–"}:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        number = float(match.group(0))
        return number if math.isfinite(number) else None
    except Exception:
        return None


def _volume_ml(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().lower().replace(" ", "")
    if not text or text in {"-", "—", "–"}:
        return None
    # summed forms such as 800ml+300ml are common in the historical workbook
    parts = re.findall(r"(\d+(?:\.\d+)?)(ml|l)", text)
    if not parts:
        return _float(text)
    total = 0.0
    for number, unit in parts:
        total += float(number) * (1000.0 if unit == "l" else 1.0)
    return total


def _scale_mmol(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).lower().replace(" ", "")
    number = _float(text)
    if number is None:
        return None
    # Some legacy rows say ml although the report is clearly using synthesis scale.
    return number


def _eq_time(value: Any) -> tuple[float | None, float | None]:
    text = str(value or "").lower().replace(" ", "")
    eq_match = re.search(r"(\d+(?:\.\d+)?)eq", text)
    h_match = re.search(r"(\d+(?:\.\d+)?)h", text)
    return (
        float(eq_match.group(1)) if eq_match else None,
        float(h_match.group(1)) if h_match else None,
    )


def normalize_resin(value: Any) -> str:
    text = str(value or "").strip()
    low = text.lower()
    if "wang" in low:
        return "Wang resin"
    if "rink" in low and "amide" in low:
        return "Rink Amide resin"
    if "trityl" in low or "ctc" in low:
        return "Trityl/2-CTC resin"
    return text


AA_ALIASES = {
    "A": "Fmoc-Ala-OH", "R": "Fmoc-Arg(Pbf)-OH", "N": "Fmoc-Asn(Trt)-OH",
    "D": "Fmoc-Asp(OtBu)-OH", "C": "Fmoc-Cys(Trt)-OH", "Q": "Fmoc-Gln(Trt)-OH",
    "E": "Fmoc-Glu(OtBu)-OH", "G": "Fmoc-Gly-OH", "H": "Fmoc-His(Trt)-OH",
    "I": "Fmoc-Ile-OH", "L": "Fmoc-Leu-OH", "K": "Fmoc-Lys(Boc)-OH",
    "M": "Fmoc-Met-OH", "F": "Fmoc-Phe-OH", "P": "Fmoc-Pro-OH",
    "S": "Fmoc-Ser(tBu)-OH", "T": "Fmoc-Thr(tBu)-OH", "W": "Fmoc-Trp(Boc)-OH",
    "Y": "Fmoc-Tyr(tBu)-OH", "V": "Fmoc-Val-OH",
}


def normalize_amino_acid(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    compact = re.sub(r"\s+", "", text)
    compact = compact.replace("Fmoc-", "Fmoc-")
    if compact in AA_ALIASES:
        return AA_ALIASES[compact]
    replacements = {
        "fmoc-arg(pbf)-oh": "Fmoc-Arg(Pbf)-OH",
        "fmoc-asn(trt)-oh": "Fmoc-Asn(Trt)-OH",
        "fmoc-asp(otbu)-oh": "Fmoc-Asp(OtBu)-OH",
        "fmoc-cys(trt)-oh": "Fmoc-Cys(Trt)-OH",
        "fmoc-cys(acm)-oh": "Fmoc-Cys(Acm)-OH",
        "fmoc-gln(trt)-oh": "Fmoc-Gln(Trt)-OH",
        "fmoc-glu(otbu)-oh": "Fmoc-Glu(OtBu)-OH",
        "fmoc-his(trt)-oh": "Fmoc-His(Trt)-OH",
        "fmoc-lys(boc)-oh": "Fmoc-Lys(Boc)-OH",
        "fmoc-ser(tbu)-oh": "Fmoc-Ser(tBu)-OH",
        "fmoc-thr(tbu)-oh": "Fmoc-Thr(tBu)-OH",
        "fmoc-trp(boc)-oh": "Fmoc-Trp(Boc)-OH",
        "fmoc-tyr(tbu)-oh": "Fmoc-Tyr(tBu)-OH",
        "fmoc-hyp(tbu)-oh": "Fmoc-Hyp(tBu)-OH",
        "fmoc-cit-oh": "Fmoc-Cit-OH",
        "fmoc-aeea-oh": "Fmoc-AEEA-OH",
    }
    return replacements.get(compact.lower(), text)


def _stereo(name: str) -> str:
    low = name.lower()
    return "D" if "fmoc-d-" in low or low.startswith("d-") else "L/unspecified"


def _protecting_group(name: str) -> str:
    match = re.search(r"\(([^)]+)\)", name)
    return match.group(1) if match else ""


def _keyword_flags(observation: str, filter_note: str = "") -> dict[str, int | None]:
    text = f"{observation}\n{filter_note}".lower()
    good_precip = any(token in text for token in ("석출 잘", "석출은 잘", "침전이 빨", "가루처럼 잘"))
    bad_precip = any(token in text for token in ("석출이 잘 안", "석출x", "침전 안", "석출 안"))
    separation_problem = any(token in text for token in ("분리가 잘 안", "상등액을 버리기 애매", "상층액을 버리기 애매", "상등액 분리 어려"))
    concentrate = any(token in text for token in ("농축 후", "농축해서", "농축후", "농축 추천"))
    remove_tis = "tis는 빼" in text or "tis 빼" in text
    hardening = any(token in text for token in ("overnight", "딱딱", "케이크처럼", "떡지"))
    return {
        "precipitation_good": 0 if bad_precip else (1 if good_precip else None),
        "separation_problem": 1 if separation_problem else 0 if "분리 잘 됨" in text else None,
        "concentration_recommended": 1 if concentrate else 0,
        "remove_tis_recommended": 1 if remove_tis else 0,
        "overnight_hardening": 1 if hardening else 0,
    }


def _crude_g(text: str) -> float | None:
    matches = re.findall(r"(?:cr(?:ude)?\s*[:=]?\s*)(\d+(?:\.\d+)?)\s*g", text, flags=re.I)
    if matches:
        return float(matches[-1])
    return None


def _insert_cleavage(con: sqlite3.Connection, row: Mapping[str, Any]) -> bool:
    existing = con.execute(
        """SELECT record_id FROM cleavage_records
           WHERE source_id=? AND source_locator=?""",
        (row.get("source_id"), row.get("source_locator", "")),
    ).fetchone()
    if existing:
        return False
    keys = [
        "record_id", "status", "product", "sequence", "scale_mmol", "operator",
        "tfa_ml", "tis_ml", "water_ml", "other_scavengers_json", "cleavage_eq",
        "cleavage_time_h", "temperature_c", "ether_ml", "ether_ratio",
        "filter_ether_ml", "filter_speed", "crude_g", "precipitation_good",
        "separation_problem", "concentration_recommended", "remove_tis_recommended",
        "overnight_hardening", "raw_observation", "raw_filter_note", "source_id",
        "source_locator", "outlier_flag", "created_at", "updated_at",
    ]
    values = [row.get(key) for key in keys]
    con.execute(
        f"INSERT INTO cleavage_records ({','.join(keys)}) VALUES ({','.join('?' for _ in keys)})",
        values,
    )
    return True


def import_cleavage_report(path: str | Path, db_path: str | Path | None = None) -> dict[str, Any]:
    from openpyxl import load_workbook
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(source)
    initialize(db_path)
    workbook = load_workbook(source, data_only=True, read_only=True)
    inserted = 0
    with _connect(db_path) as con:
        source_id = _register_source(con, source, "cleavage_report_xlsx")
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            index = 0
            while index < len(rows):
                row = rows[index]
                product = str(row[0] or "").strip() if row else ""
                # A report block starts with a product row and is followed by Cleavage/TFA.
                if product and index + 1 < len(rows):
                    next_row = rows[index + 1]
                    if str(next_row[0] or "").strip().lower() == "cleavage" and str(next_row[1] or "").strip().upper() == "TFA":
                        scale = _scale_mmol(row[2] if len(row) > 2 else None)
                        operator = str(row[3] or "").strip() if len(row) > 3 else ""
                        tfa_row = next_row
                        tfa_ml = _volume_ml(tfa_row[2] if len(tfa_row) > 2 else None)
                        cleavage_eq, cleavage_time_h = _eq_time(tfa_row[3] if len(tfa_row) > 3 else None)
                        observation = str(tfa_row[4] or "").strip() if len(tfa_row) > 4 else ""
                        components: dict[str, Any] = {}
                        ether_ml = None; ether_ratio = ""; filter_ether_ml = None; filter_speed = ""; filter_note = ""
                        j = index + 2
                        while j < min(index + 8, len(rows)):
                            current = rows[j]
                            section = str(current[0] or "").strip().lower() if current else ""
                            name = str(current[1] or "").strip() if len(current) > 1 else ""
                            value = current[2] if len(current) > 2 else None
                            detail = str(current[3] or "").strip() if len(current) > 3 else ""
                            note = str(current[4] or "").strip() if len(current) > 4 else ""
                            if not any(v not in (None, "") for v in current[:5]):
                                break
                            if section == "filter":
                                filter_ether_ml = _volume_ml(value)
                                filter_speed = detail
                                filter_note = note
                            elif name.upper() == "TIS":
                                components["TIS"] = _volume_ml(value)
                            elif name.upper() in {"H2O", "WATER"}:
                                components["H2O"] = _volume_ml(value)
                            elif name.lower() == "ether":
                                ether_ml = _volume_ml(value)
                                ether_ratio = detail
                            elif name:
                                components[name] = value
                            j += 1
                        flags = _keyword_flags(observation, filter_note)
                        record = {
                            "record_id": _id(), "status": "parsed", "product": product,
                            "sequence": "", "scale_mmol": scale, "operator": operator,
                            "tfa_ml": tfa_ml, "tis_ml": components.pop("TIS", None),
                            "water_ml": components.pop("H2O", None),
                            "other_scavengers_json": json.dumps(components, ensure_ascii=False, default=str),
                            "cleavage_eq": cleavage_eq, "cleavage_time_h": cleavage_time_h,
                            "temperature_c": None, "ether_ml": ether_ml, "ether_ratio": ether_ratio,
                            "filter_ether_ml": filter_ether_ml, "filter_speed": filter_speed,
                            "crude_g": _crude_g(f"{observation}\n{filter_note}"),
                            **flags,
                            "raw_observation": observation, "raw_filter_note": filter_note,
                            "source_id": source_id, "source_locator": f"{sheet.title}!row{index + 1}",
                            "outlier_flag": 0, "created_at": _now(), "updated_at": _now(),
                        }
                        inserted += int(_insert_cleavage(con, record))
                        index = j
                        continue
                index += 1
    return {"kind": "cleavage", "inserted": inserted, "source": str(source)}


def _mapping_value(row: Mapping[str, Any], *names: str) -> Any:
    lookup = {str(key).strip().lower(): value for key, value in row.items()}
    for name in names:
        if name.lower() in lookup:
            return lookup[name.lower()]
    return None


def _insert_loading(con: sqlite3.Connection, row: Mapping[str, Any]) -> bool:
    existing = con.execute(
        "SELECT record_id FROM loading_records WHERE source_id=? AND source_locator=?",
        (row.get("source_id"), row.get("source_locator", "")),
    ).fetchone()
    if existing:
        return False
    keys = [
        "record_id", "status", "date", "resin_type", "resin_note", "amino_acid_raw",
        "amino_acid_normalized", "stereochemistry", "protecting_group", "aa_eq", "base",
        "base_eq", "coupling_reagent", "coupling_reagent_eq", "additive", "additive_eq",
        "loading_time_h", "loading_solvent", "capping_performed", "capping_method",
        "resin_sample_weight_mg", "absorbance", "loading_rate_mmol_g", "raw_note",
        "source_id", "source_locator", "outlier_flag", "created_at", "updated_at",
    ]
    con.execute(
        f"INSERT INTO loading_records ({','.join(keys)}) VALUES ({','.join('?' for _ in keys)})",
        [row.get(key) for key in keys],
    )
    return True


def import_loading_csv(path: str | Path, db_path: str | Path | None = None) -> dict[str, Any]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(source)
    initialize(db_path)
    inserted = 0
    with source.open("r", encoding="utf-8-sig", newline="") as handle, _connect(db_path) as con:
        reader = csv.DictReader(handle)
        source_id = _register_source(con, source, "loading_csv")
        for number, raw in enumerate(reader, 2):
            aa_raw = str(_mapping_value(raw, "amino_acid", "amino_acid_raw", "aa") or "").strip()
            aa = normalize_amino_acid(_mapping_value(raw, "amino_acid_normalized") or aa_raw)
            resin_raw = _mapping_value(raw, "resin_type", "resin")
            note = str(_mapping_value(raw, "raw_note", "note", "비고") or "")
            capping_raw = _mapping_value(raw, "capping_performed", "capping")
            capping = None
            if capping_raw is not None and str(capping_raw).strip() != "":
                capping = 1 if str(capping_raw).strip().lower() in {"1", "true", "yes", "y", "있음", "약식"} else 0
            record = {
                "record_id": _id(), "status": str(_mapping_value(raw, "status") or "parsed").lower(),
                "date": str(_mapping_value(raw, "date", "날짜") or ""),
                "resin_type": normalize_resin(resin_raw), "resin_note": str(resin_raw or ""),
                "amino_acid_raw": aa_raw, "amino_acid_normalized": aa,
                "stereochemistry": _stereo(aa), "protecting_group": _protecting_group(aa),
                "aa_eq": _float(_mapping_value(raw, "aa_eq", "loading_aa_eq")),
                "base": str(_mapping_value(raw, "base") or "DIEA"),
                "base_eq": _float(_mapping_value(raw, "base_eq", "diea_eq")),
                "coupling_reagent": str(_mapping_value(raw, "coupling_reagent") or ""),
                "coupling_reagent_eq": _float(_mapping_value(raw, "coupling_reagent_eq")),
                "additive": str(_mapping_value(raw, "additive") or ""),
                "additive_eq": _float(_mapping_value(raw, "additive_eq")),
                "loading_time_h": _float(_mapping_value(raw, "loading_time_h", "time_h")),
                "loading_solvent": str(_mapping_value(raw, "loading_solvent", "solvent") or ""),
                "capping_performed": capping,
                "capping_method": str(_mapping_value(raw, "capping_method") or ""),
                "resin_sample_weight_mg": _float(_mapping_value(raw, "resin_sample_weight_mg", "resin_weight_mg")),
                "absorbance": _float(_mapping_value(raw, "absorbance", "abs")),
                "loading_rate_mmol_g": _float(_mapping_value(raw, "loading_rate_mmol_g", "loading_rate")),
                "raw_note": note, "source_id": source_id, "source_locator": f"row{number}",
                "outlier_flag": int(str(_mapping_value(raw, "outlier_flag") or "0").strip() in {"1", "true", "True"}),
                "created_at": _now(), "updated_at": _now(),
            }
            if record["status"] not in STATUSES:
                record["status"] = "parsed"
            inserted += int(_insert_loading(con, record))
    flag_loading_outliers(db_path)
    return {"kind": "loading", "inserted": inserted, "source": str(source)}


def import_path(path: str | Path, db_path: str | Path | None = None) -> list[dict[str, Any]]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(source)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return [import_loading_csv(source, db_path)]
    if suffix in {".xlsx", ".xlsm"}:
        # Only Cleavage Report has a stable schema today. Other workbooks are registered
        # without fabricating structure; they remain available for later reviewed parsing.
        try:
            from openpyxl import load_workbook
            wb = load_workbook(source, data_only=True, read_only=True)
            first = wb[wb.sheetnames[0]]
            first_values = [str(cell.value or "") for row in first.iter_rows(min_row=1, max_row=10) for cell in row]
            looks_like_cleavage = any("Cleavage" in value for value in first_values) and any("TFA" in value for value in first_values)
            wb.close()
        except Exception:
            looks_like_cleavage = False
        if looks_like_cleavage:
            return [import_cleavage_report(source, db_path)]
        initialize(db_path)
        with _connect(db_path) as con:
            _register_source(con, source, "historical_workbook", note="Registered for reviewed V4 parsing; no guessed rows were created.")
        return [{"kind": "registered_workbook", "inserted": 0, "source": str(source)}]
    if suffix == ".zip":
        results: list[dict[str, Any]] = []
        with tempfile.TemporaryDirectory(prefix="spps_v4_import_") as temp:
            with zipfile.ZipFile(source) as archive:
                for info in archive.infolist():
                    name = Path(info.filename)
                    if info.is_dir() or name.name.startswith("~$") or name.suffix.lower() not in {".xlsx", ".xlsm", ".csv"}:
                        continue
                    safe = Path(temp) / f"{len(results):04d}_{name.name}"
                    safe.write_bytes(archive.read(info))
                    results.extend(import_path(safe, db_path))
        return results
    raise ValueError(f"Unsupported experimental data file: {source.suffix}")


def list_records(kind: str, db_path: str | Path | None = None, *, statuses: Iterable[str] | None = None) -> list[dict[str, Any]]:
    initialize(db_path)
    table = "loading_records" if kind == "loading" else "cleavage_records" if kind == "cleavage" else None
    if table is None:
        raise ValueError("kind must be loading or cleavage")
    clauses = ""
    params: list[Any] = []
    if statuses:
        values = [value for value in statuses if value in STATUSES]
        if values:
            clauses = f" WHERE status IN ({','.join('?' for _ in values)})"
            params.extend(values)
    with _connect(db_path) as con:
        rows = con.execute(f"SELECT * FROM {table}{clauses} ORDER BY created_at, record_id", params).fetchall()
    return [dict(row) for row in rows]


def set_status(kind: str, record_ids: Iterable[str], status: str, db_path: str | Path | None = None) -> int:
    if status not in STATUSES:
        raise ValueError(f"Unsupported status: {status}")
    table = "loading_records" if kind == "loading" else "cleavage_records" if kind == "cleavage" else None
    if table is None:
        raise ValueError("kind must be loading or cleavage")
    ids = [str(value) for value in record_ids if str(value)]
    if not ids:
        return 0
    with _connect(db_path) as con:
        cur = con.execute(
            f"UPDATE {table} SET status=?, updated_at=? WHERE record_id IN ({','.join('?' for _ in ids)})",
            [status, _now(), *ids],
        )
        return int(cur.rowcount)



def update_record(kind: str, record_id: str, changes: Mapping[str, Any], db_path: str | Path | None = None) -> dict[str, Any]:
    table = "loading_records" if kind == "loading" else "cleavage_records" if kind == "cleavage" else None
    if table is None:
        raise ValueError("kind must be loading or cleavage")
    editable = {
        "loading": {"status", "date", "resin_type", "resin_note", "amino_acid_raw", "amino_acid_normalized", "aa_eq", "base", "base_eq", "coupling_reagent", "coupling_reagent_eq", "additive", "additive_eq", "loading_time_h", "loading_solvent", "capping_performed", "capping_method", "resin_sample_weight_mg", "absorbance", "loading_rate_mmol_g", "raw_note", "outlier_flag"},
        "cleavage": {"status", "product", "sequence", "scale_mmol", "operator", "tfa_ml", "tis_ml", "water_ml", "other_scavengers_json", "cleavage_eq", "cleavage_time_h", "temperature_c", "ether_ml", "ether_ratio", "filter_ether_ml", "filter_speed", "crude_g", "precipitation_good", "separation_problem", "concentration_recommended", "remove_tis_recommended", "overnight_hardening", "raw_observation", "raw_filter_note", "outlier_flag"},
    }[kind]
    clean = {str(key): value for key, value in changes.items() if str(key) in editable}
    if "status" in clean and clean["status"] not in STATUSES:
        raise ValueError(f"Unsupported status: {clean['status']}")
    if not clean:
        raise ValueError("No editable fields were supplied.")
    clean["updated_at"] = _now()
    with _connect(db_path) as con:
        found = con.execute(f"SELECT record_id FROM {table} WHERE record_id=?", (record_id,)).fetchone()
        if not found:
            raise ValueError("Experimental record was not found.")
        assignments = ",".join(f"{key}=?" for key in clean)
        con.execute(f"UPDATE {table} SET {assignments} WHERE record_id=?", [*clean.values(), record_id])
        row = con.execute(f"SELECT * FROM {table} WHERE record_id=?", (record_id,)).fetchone()
    return dict(row)


def add_record(kind: str, values: Mapping[str, Any], db_path: str | Path | None = None, *, status: str = "verified") -> dict[str, Any]:
    """Insert one operator-entered experimental record without fabricating missing values."""
    if status not in STATUSES:
        raise ValueError(f"Unsupported status: {status}")
    initialize(db_path)
    table = "loading_records" if kind == "loading" else "cleavage_records" if kind == "cleavage" else None
    if table is None:
        raise ValueError("kind must be loading or cleavage")
    editable = {
        "loading": {"date", "resin_type", "resin_note", "amino_acid_raw", "amino_acid_normalized", "stereochemistry", "protecting_group", "aa_eq", "base", "base_eq", "coupling_reagent", "coupling_reagent_eq", "additive", "additive_eq", "loading_time_h", "loading_solvent", "capping_performed", "capping_method", "resin_sample_weight_mg", "absorbance", "loading_rate_mmol_g", "raw_note", "outlier_flag"},
        "cleavage": {"product", "sequence", "scale_mmol", "operator", "tfa_ml", "tis_ml", "water_ml", "other_scavengers_json", "cleavage_eq", "cleavage_time_h", "temperature_c", "ether_ml", "ether_ratio", "filter_ether_ml", "filter_speed", "crude_g", "precipitation_good", "separation_problem", "concentration_recommended", "remove_tis_recommended", "overnight_hardening", "raw_observation", "raw_filter_note", "outlier_flag"},
    }[kind]
    now = _now()
    row = {str(k): v for k, v in values.items() if str(k) in editable}
    row.update({"record_id": _id(), "status": status, "created_at": now, "updated_at": now})
    if kind == "loading":
        raw = str(row.get("amino_acid_raw") or row.get("amino_acid_normalized") or "").strip()
        if raw and not row.get("amino_acid_normalized"):
            row["amino_acid_normalized"] = normalize_amino_acid(raw)
        name = str(row.get("amino_acid_normalized") or raw)
        row.setdefault("stereochemistry", _stereo(name))
        row.setdefault("protecting_group", _protecting_group(name))
        if row.get("resin_type"):
            row["resin_type"] = normalize_resin(row["resin_type"])
    keys = list(row)
    with _connect(db_path) as con:
        con.execute(f"INSERT INTO {table} ({','.join(keys)}) VALUES ({','.join('?' for _ in keys)})", [row[k] for k in keys])
        saved = con.execute(f"SELECT * FROM {table} WHERE record_id=?", (row["record_id"],)).fetchone()
    return dict(saved)

def flag_loading_outliers(db_path: str | Path | None = None) -> int:
    """Flag statistical/physical review candidates without deleting any value."""
    rows = list_records("loading", db_path)
    changed = 0
    with _connect(db_path) as con:
        for row in rows:
            value = row.get("loading_rate_mmol_g")
            flag = 1 if value is not None and (float(value) < 0 or float(value) > 2.5) else 0
            # Absorbance/loading transcription mistakes often create implausible >2.5 mmol/g values.
            if flag != int(row.get("outlier_flag") or 0):
                con.execute("UPDATE loading_records SET outlier_flag=?, updated_at=? WHERE record_id=?", (flag, _now(), row["record_id"]))
                changed += 1
    return changed


def sources(db_path: str | Path | None = None) -> list[dict[str, Any]]:
    initialize(db_path)
    with _connect(db_path) as con:
        rows = con.execute("SELECT * FROM import_sources ORDER BY imported_at DESC").fetchall()
    return [dict(row) for row in rows]


__all__ = [
    "SCHEMA_VERSION", "STATUSES", "default_db_path", "initialize", "import_cleavage_report",
    "import_loading_csv", "import_path", "list_records", "normalize_amino_acid", "normalize_resin",
    "set_status", "update_record", "add_record", "sources", "flag_loading_outliers",
]
